# -*- coding: utf-8 -*-
"""
mgt2excel — midas iGen の MGT/ANL ファイルから断面算定用 Excel を作成するツール

機能:
  - MGT(モデル) と ANL(解析結果テキスト) を読み取り、4シートの Excel を出力
      ①部材一覧   … 分割要素を一部材に統合。通し長さ=座屈長さ、LC別断面力 max/min
      ②要素データ … 全要素×荷重組合せ×I/J端の断面力(部材IDで①と紐付け)
      ③断面リスト ④節点座標
  - 部材統合の基準: 直線・同断面・同材料・同タイプで節点連鎖し、
      共有節点に 梁端解放(ピン等) / 支点 が無いこと(いずれもオプションで変更可)。
      同断面が3本以上集まる節点では統合しない(曖昧なため)
  - 単位系は ANL/MGT から自動判定し kN・kN·m・mm に換算(換算なしも選択可)

使い方:
  GUI : 引数なしで起動(exe はダブルクリック)
  CLI : py -3 mgt2excel.py model.mgt [model.anl] [-o out.xlsx]
            [--no-support-split] [--no-release-split] [--no-unit-convert]
"""

import argparse
import datetime
import math
import os
import re
import sys
import threading
import traceback
from collections import defaultdict

VERSION = "1.0.0"

FORCE_TO_KN = {"TONF": 9.80665, "KGF": 9.80665e-3, "KN": 1.0, "N": 1e-3}
LEN_TO_MM = {"MM": 1.0, "CM": 10.0, "M": 1000.0}
COS_TOL = 0.9999  # 約0.8度以内を同一直線とみなす
COMPS = ["N", "Qy", "Qz", "Mt", "My", "Mz"]
LINE_TYPES = ("BEAM", "TRUSS", "TENSTR", "COMPTR")  # 出力対象の線要素


def _print(msg):
    """コンソールの文字コードで表せない文字があっても落ちない print"""
    try:
        print(msg)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "utf-8"
        print(str(msg).encode(enc, "replace").decode(enc))


# ---------------------------------------------------------------- MGT parse
def _read_lines(path):
    with open(path, encoding="cp932", errors="replace") as f:
        return f.read().splitlines()


def _split_sections(lines):
    sec = {}
    cur = None
    for ln in lines:
        if ln.startswith("*"):
            cur = ln.split(";")[0].strip().upper()
            sec.setdefault(cur, [])
        elif cur is not None:
            sec[cur].append(ln)
    return sec


def _data_lines(block):
    return [l for l in block if l.strip() and not l.lstrip().startswith(";")]


_RE_FLAG6 = re.compile(r"[01]{6}")


def _expand_list(tok):
    """MGT の節点/要素リスト表記 '5' '5to10' '5to10by2' を展開する"""
    m = re.match(r"^(\d+)(?:to(\d+)(?:by(\d+))?)?$", tok.strip())
    if not m:
        return []
    a = int(m.group(1))
    b = int(m.group(2)) if m.group(2) else a
    s = int(m.group(3)) if m.group(3) else 1
    return list(range(a, b + 1, s))


def parse_mgt(path, log=_print):
    mgt = _split_sections(_read_lines(path))
    for req in ("*NODE", "*ELEMENT"):
        if req not in mgt:
            raise ValueError("MGT に {} セクションがありません: {}".format(req, path))

    unit_force, unit_len = "TONF", "MM"
    if mgt.get("*UNIT"):
        f = [t.strip().upper() for t in _data_lines(mgt["*UNIT"])[0].split(",")]
        unit_force, unit_len = f[0], f[1]

    nodes = {}
    for l in _data_lines(mgt["*NODE"]):
        f = [t.strip() for t in l.split(",")]
        nodes[int(f[0])] = (float(f[1]), float(f[2]), float(f[3]))

    elements = {}
    for l in _data_lines(mgt["*ELEMENT"]):
        f = [t.strip() for t in l.split(",")]
        eid, etype = int(f[0]), f[1].upper()
        if etype in LINE_TYPES:
            n1, n2 = int(f[4]), int(f[5])
            try:
                a, b = nodes[n1], nodes[n2]
            except KeyError as ke:
                raise ValueError("要素 {} が存在しない節点 {} を参照しています"
                                 .format(eid, ke.args[0])) from None
            v = (b[0] - a[0], b[1] - a[1], b[2] - a[2])
            ln_ = math.sqrt(sum(c * c for c in v))
            if ln_ == 0:
                log("警告: 長さ0の要素 {} をスキップしました".format(eid))
                continue
            raw_beta = f[6] if len(f) > 6 and f[6] else "0"
            try:
                beta = float(raw_beta)  # "0" と "0.0" の表記ゆれを吸収
            except ValueError:
                beta = raw_beta
            elements[eid] = dict(id=eid, type=etype, mat=int(f[2]), sec=int(f[3]),
                                 n1=n1, n2=n2, beta=beta,
                                 vec=tuple(c / ln_ for c in v), len=ln_)

    materials = {}
    for l in _data_lines(mgt.get("*MATERIAL", [])):
        f = [t.strip() for t in l.split(",")]
        if f[0].isdigit():
            materials[int(f[0])] = f[2]

    sections = {}
    for l in _data_lines(mgt.get("*SECTION", [])):
        f = [t.strip() for t in l.split(",")]
        if not f[0].isdigit():
            continue  # 継続行(2行目以降)は無視
        sections[int(f[0])] = dict(id=int(f[0]), type=f[1], name=f[2],
                                   shape=f[12] if len(f) > 12 else "",
                                   data=", ".join(f[13:]) if len(f) > 13 else "")

    # FRAME-RLS: 梁端解放。2行1組 (I端フラグ行 + J端フラグ行)、フラグは Fx..Mz の6桁。
    # フラグ形式を検証し、ペアが崩れた行は警告してスキップする(サイレントなずれを防ぐ)
    releases = {}
    rls = _data_lines(mgt.get("*FRAME-RLS", []))
    i = 0
    while i < len(rls):
        f1 = [t.strip() for t in rls[i].split(",")]
        f2 = [t.strip() for t in rls[i + 1].split(",")] if i + 1 < len(rls) else []
        if (len(f1) >= 3 and _RE_FLAG6.fullmatch(f1[2])
                and f2 and _RE_FLAG6.fullmatch(f2[0])):
            for tok in f1[0].split():
                for e in _expand_list(tok):
                    releases[e] = (f1[2], f2[0])
            i += 2
        else:
            log("警告: *FRAME-RLS を解釈できない行をスキップしました: "
                + rls[i].strip()[:40])
            i += 1

    supports = set()
    for l in _data_lines(mgt.get("*CONSTRAINT", [])):
        for tok in l.split(",")[0].split():
            supports.update(_expand_list(tok))

    log("MGT 読込: 節点 {} / 線要素 {} / 断面 {} / 梁端解放 {} / 支点 {}".format(
        len(nodes), len(elements), len(sections), len(releases), len(supports)))
    return dict(nodes=nodes, elements=elements, materials=materials,
                sections=sections, releases=releases, supports=supports,
                unit_force=unit_force, unit_len=unit_len)


# ---------------------------------------------------------------- ANL parse
_FLOAT6 = r"((?:\s+-?\d+\.\d+){6})\s*$"
_PT = r"([IJ]|CNT|\d+/\d+)"  # 出力位置: I / J / 中央(CNT) / 中間点(1/4, 3/4 等)
_RE_BELEM = re.compile(r"^\s*(\d+)\s+(\d+)\s+(\d+)\s+(\S+)\s+" + _PT + _FLOAT6)
_RE_BLC = re.compile(r"^\s+(\S+)\s+" + _PT + _FLOAT6)
_RE_BPT = re.compile(r"^\s+" + _PT + _FLOAT6)
_FLOAT2 = r"(-?\d+\.\d+)\s+(-?\d+\.\d+)\s*$"
_RE_TELEM = re.compile(r"^\s*(\d+)\s+(\d+)\s+(\d+)\s+(\S+)\s+" + _FLOAT2)
_RE_TLC = re.compile(r"^\s+(\S+)\s+" + _FLOAT2)
_RE_UNIT = re.compile(r"単位系\s*:\s*(\S+)\s*,\s*(\S+)")
_RE_TAIL6 = re.compile(r"(?:-?\d+\.\d+\s+){5}-?\d+\.\d+\s*$")


def _pt_order(pt):
    """出力位置の並び順: I -> 1/4 -> CNT -> 3/4 -> J"""
    if pt == "I":
        return -1.0
    if pt == "J":
        return 2.0
    if pt == "CNT":
        return 0.5
    m = re.match(r"^(\d+)/(\d+)$", pt)
    return int(m.group(1)) / int(m.group(2)) if m else 1.5


def parse_anl(path, log=_print):
    anl = _read_lines(path)

    def find(pat, start=0):
        for i in range(start, len(anl)):
            if pat in anl[i]:
                return i
        return -1

    unit_force, unit_len = None, None
    beam_forces = defaultdict(dict)  # eid -> {(lc, "I"/"J"): [N,Qy,Qz,Mt,My,Mz]}
    b0 = find("梁要素の断面力の出力")
    if b0 >= 0:
        m = _RE_UNIT.search(anl[b0])
        if m:
            unit_force, unit_len = m.group(1).upper(), m.group(2).upper()
        b1 = find("梁要素の断面力 ---", b0)
        cur_eid = cur_lc = None
        n_unmatched = 0
        for ln in anl[b0:b1 if b1 > 0 else len(anl)]:
            m = _RE_BELEM.match(ln)
            if m:
                cur_eid, cur_lc = int(m.group(1)), m.group(4)
                beam_forces[cur_eid][(cur_lc, m.group(5))] = \
                    [float(x) for x in m.group(6).split()]
                continue
            m = _RE_BLC.match(ln)
            if m and cur_eid is not None:
                cur_lc = m.group(1)
                beam_forces[cur_eid][(cur_lc, m.group(2))] = \
                    [float(x) for x in m.group(3).split()]
                continue
            m = _RE_BPT.match(ln)
            if m and cur_eid is not None and cur_lc is not None:
                beam_forces[cur_eid][(cur_lc, m.group(1))] = \
                    [float(x) for x in m.group(2).split()]
                continue
            if _RE_TAIL6.search(ln):
                n_unmatched += 1  # 位置ラベルを解釈できなかった断面力行
        if n_unmatched:
            log("警告: 梁断面力の {} 行は出力位置を解釈できず"
                "集計対象外です".format(n_unmatched))

    truss_forces = defaultdict(dict)  # eid -> {lc: (Ni, Nj)}
    t0 = find("トラス要素の断面力の出力")
    if t0 >= 0:
        if unit_force is None:
            m = _RE_UNIT.search(anl[t0])
            if m:
                unit_force, unit_len = m.group(1).upper(), m.group(2).upper()
        t1 = find("トラス要素の応力度の出力", t0)
        cur_eid = None
        for ln in anl[t0:t1 if t1 > 0 else len(anl)]:
            m = _RE_TELEM.match(ln)
            if m:
                cur_eid = int(m.group(1))
                truss_forces[cur_eid][m.group(4)] = \
                    (float(m.group(5)), float(m.group(6)))
                continue
            m = _RE_TLC.match(ln)
            if m and cur_eid is not None:
                truss_forces[cur_eid][m.group(1)] = \
                    (float(m.group(2)), float(m.group(3)))

    if not beam_forces and not truss_forces:
        raise ValueError("ANL から断面力を読み取れませんでした。"
                         "「梁要素の断面力の出力」を含むテキスト出力か確認してください: " + path)

    lcs = []
    for d in list(beam_forces.values()) + list(truss_forces.values()):
        for k in d:
            lc = k[0] if isinstance(k, tuple) else k
            if lc not in lcs:
                lcs.append(lc)

    log("ANL 読込: 梁要素 {} / トラス要素 {} / 荷重ケース {}".format(
        len(beam_forces), len(truss_forces), ", ".join(lcs)))
    return dict(beam_forces=beam_forces, truss_forces=truss_forces, lcs=lcs,
                unit_force=unit_force, unit_len=unit_len)


# ---------------------------------------------------------------- 部材統合
def merge_members(model, split_at_supports=True, split_at_releases=True, log=_print):
    elements = model["elements"]
    releases = model["releases"]
    supports = model["supports"]

    parent = {eid: eid for eid in elements}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def released(eid, end):  # end: 0=I(n1) 1=J(n2)
        r = releases.get(eid)
        return r is not None and any(c != "0" for c in r[end])

    incid = defaultdict(list)  # 節点 -> [(eid, end)]
    for e in elements.values():
        incid[e["n1"]].append((e["id"], 0))
        incid[e["n2"]].append((e["id"], 1))

    for nid, lst in incid.items():
        if split_at_supports and nid in supports:
            continue  # 支点節点では部材を切る
        bykey = defaultdict(list)
        for eid, end in lst:
            e = elements[eid]
            bykey[(e["type"], e["mat"], e["sec"], e["beta"])].append((eid, end))
        for grp in bykey.values():
            if len(grp) != 2:
                continue  # 同断面が3本以上集まる節点は統合しない(曖昧)
            (e1, end1), (e2, end2) = grp
            if e1 == e2:
                continue
            if split_at_releases and (released(e1, end1) or released(e2, end2)):
                continue  # ピン等 -> 部材境界
            v1, v2 = elements[e1]["vec"], elements[e2]["vec"]
            # 節点を通過する向きに揃える: J側(end=1)はそのまま、I側(end=0)は反転
            d1 = v1 if end1 == 1 else tuple(-c for c in v1)
            d2 = v2 if end2 == 0 else tuple(-c for c in v2)
            if sum(a * b for a, b in zip(d1, d2)) > COS_TOL:
                ra, rb = find(e1), find(e2)
                if ra != rb:
                    parent[rb] = ra

    members = defaultdict(list)
    for eid in elements:
        members[find(eid)].append(eid)
    groups = sorted(members.values(), key=min)
    n_multi = sum(1 for g in groups if len(g) > 1)
    log("部材統合: {} 要素 -> {} 部材 (複数要素の統合 {})".format(
        len(elements), len(groups), n_multi))
    return groups


# ---------------------------------------------------------------- 集計
def _dummy_tag(e, model):
    tags = []
    mn = model["materials"].get(e["mat"], "")
    sn = model["sections"].get(e["sec"], {}).get("name", "")
    if any(k in mn.lower() for k in ("rigid", "dammy", "dummy")) or "ダミー" in mn:
        tags.append("材料:" + mn)
    if any(k in sn.lower() for k in ("dammy", "dummy")) or "ダミー" in sn:
        tags.append("断面:" + sn)
    return "、".join(tags)


def build_members(model, results, groups, convert_units=True, log=_print):
    nodes, elements = model["nodes"], model["elements"]
    beam_forces, truss_forces = results["beam_forces"], results["truss_forces"]
    lcs = results["lcs"]

    # 断面力(uf, ul_anl)は ANL の単位系、幾何量=座標・長さ(ul_geo)は MGT の単位系で換算する
    uf = (results["unit_force"] or model["unit_force"]).upper()
    ul_anl = (results["unit_len"] or model["unit_len"]).upper()
    ul_geo = model["unit_len"].upper()
    if ul_anl != ul_geo:
        log("注意: 長さ単位が MGT({}) と ANL({}) で異なります。"
            "幾何量は MGT、モーメントは ANL の単位で換算します".format(ul_geo, ul_anl))
    if convert_units:
        if uf not in FORCE_TO_KN or ul_anl not in LEN_TO_MM or ul_geo not in LEN_TO_MM:
            raise ValueError("未対応の単位系です: {} , {} / {}".format(uf, ul_anl, ul_geo))
        f_force = FORCE_TO_KN[uf]
        f_mom = FORCE_TO_KN[uf] * LEN_TO_MM[ul_anl] / 1000.0  # -> kN·m
        f_len = LEN_TO_MM[ul_geo]
        u_force, u_mom, u_len = "kN", "kN·m", "mm"
    else:
        f_force = f_mom = f_len = 1.0
        u_force, u_mom, u_len = uf, "{}·{}".format(uf, ul_anl), ul_geo
    units = dict(force=u_force, mom=u_mom, len=u_len,
                 f_force=f_force, f_mom=f_mom, f_len=f_len)
    log("単位: 断面力 {} , {} / 幾何 {} -> {} / {} / {}".format(
        uf, ul_anl, ul_geo, u_force, u_mom, u_len))

    rows = []
    elem_to_member = {}
    for idx, eids in enumerate(groups, 1):
        mid = "M{:04d}".format(idx)
        for eid in eids:
            elem_to_member[eid] = mid
        es = [elements[e] for e in eids]
        e0 = es[0]
        cnt = defaultdict(int)
        for e in es:
            cnt[e["n1"]] += 1
            cnt[e["n2"]] += 1
        ends = [n for n, c in cnt.items() if c == 1]
        na, nb = sorted(ends) if len(ends) == 2 else (e0["n1"], e0["n2"])
        length = sum(e["len"] for e in es) * f_len

        agg = {}
        for lc in lcs:
            vals = defaultdict(list)
            for e in es:
                if e["type"] == "BEAM":
                    for (l, _pt), v in beam_forces.get(e["id"], {}).items():
                        if l == lc:
                            for c, x in zip(COMPS, v):
                                vals[c].append(x * (f_force if c in ("N", "Qy", "Qz") else f_mom))
                else:
                    fij = truss_forces.get(e["id"], {}).get(lc)
                    if fij:
                        vals["N"].extend([fij[0] * f_force, fij[1] * f_force])
            for c in COMPS:
                if vals[c]:
                    agg[(lc, c, "max")] = max(vals[c])
                    agg[(lc, c, "min")] = min(vals[c])

        pa, pb = nodes[na], nodes[nb]
        rows.append(dict(
            mid=mid, type=e0["type"], sec_id=e0["sec"],
            sec=model["sections"].get(e0["sec"], {}).get("name", str(e0["sec"])),
            mat=model["materials"].get(e0["mat"], str(e0["mat"])),
            n_elem=len(eids), eids=", ".join(str(i) for i in sorted(eids)),
            na=na, nb=nb,
            pa=tuple(c * f_len for c in pa), pb=tuple(c * f_len for c in pb),
            length=length, dummy=_dummy_tag(e0, model), agg=agg))
    return rows, elem_to_member, units


# ---------------------------------------------------------------- Excel
def write_excel(out_path, model, results, rows, elem_to_member, units, log=_print):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    lcs = results["lcs"]
    hd = Font(bold=True, size=9)
    fill = PatternFill("solid", fgColor="DDEBF7")

    def style_header(ws):
        for c in ws[1]:
            c.font = hd
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "部材一覧"
    ul, uf, um = units["len"], units["force"], units["mom"]
    head = ["部材ID", "タイプ", "断面ID", "断面名", "材料", "要素数", "構成要素ID",
            "節点i", "節点j", "Xi", "Yi", "Zi", "Xj", "Yj", "Zj",
            "部材長さL=座屈長さ({})".format(ul), "ダミー", "LC"]
    for c in COMPS:
        u = uf if c in ("N", "Qy", "Qz") else um
        head += ["{}_max({})".format(c, u), "{}_min({})".format(c, u)]
    ws.append(head)
    for r in rows:
        base = [r["mid"], r["type"], r["sec_id"], r["sec"], r["mat"],
                r["n_elem"], r["eids"], r["na"], r["nb"],
                round(r["pa"][0], 1), round(r["pa"][1], 1), round(r["pa"][2], 1),
                round(r["pb"][0], 1), round(r["pb"][1], 1), round(r["pb"][2], 1),
                round(r["length"], 1), r["dummy"]]
        for lc in lcs:  # 荷重ケースごとに1行
            row = base + [lc]
            for c in COMPS:
                for mm in ("max", "min"):
                    v = r["agg"].get((lc, c, mm))
                    row.append(round(v, 3) if v is not None else None)
            ws.append(row)
    style_header(ws)
    ws.column_dimensions["G"].width = 30

    ws = wb.create_sheet("要素データ")
    ws.append(["要素ID", "部材ID", "タイプ", "断面名", "材料", "節点i", "節点j",
               "長さ({})".format(ul), "LC", "端",
               "N({})".format(uf), "Qy({})".format(uf), "Qz({})".format(uf),
               "Mt({})".format(um), "My({})".format(um), "Mz({})".format(um)])
    f_force, f_mom, f_len = units["f_force"], units["f_mom"], units["f_len"]
    for eid in sorted(model["elements"]):
        e = model["elements"][eid]
        base = [eid, elem_to_member.get(eid, ""), e["type"],
                model["sections"].get(e["sec"], {}).get("name", ""),
                model["materials"].get(e["mat"], ""),
                e["n1"], e["n2"], round(e["len"] * f_len, 1)]
        if e["type"] == "BEAM":
            d = results["beam_forces"].get(eid, {})
            for lc in lcs:
                pts = sorted({p for (l, p) in d if l == lc}, key=_pt_order)
                for pt in pts:
                    v = d[(lc, pt)]
                    conv = [round(v[i] * (f_force if i < 3 else f_mom), 3)
                            for i in range(6)]
                    ws.append(base + [lc, pt] + conv)
        else:
            d = results["truss_forces"].get(eid, {})
            for lc in lcs:
                fij = d.get(lc)
                if fij is None:
                    continue
                for pt, val in zip(("I", "J"), fij):
                    ws.append(base + [lc, pt, round(val * f_force, 3),
                                      None, None, None, None, None])
    style_header(ws)

    ws = wb.create_sheet("断面リスト")
    ws.append(["断面ID", "定義タイプ", "断面名", "形状", "寸法・定義データ"])
    for sid in sorted(model["sections"]):
        s = model["sections"][sid]
        ws.append([s["id"], s["type"], s["name"], s["shape"], s["data"]])
    style_header(ws)
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["E"].width = 60

    ws = wb.create_sheet("節点座標")
    ws.append(["節点ID", "X({})".format(ul), "Y({})".format(ul), "Z({})".format(ul)])
    for nid in sorted(model["nodes"]):
        x, y, z = model["nodes"][nid]
        ws.append([nid, round(x * f_len, 1), round(y * f_len, 1),
                   round(z * f_len, 1)])
    style_header(ws)

    # 保存。ロックされていたらタイムスタンプ付きの別名で保存する
    try:
        wb.save(out_path)
    except PermissionError:
        stem, ext = os.path.splitext(out_path)
        alt = "{}_{}{}".format(stem, datetime.datetime.now().strftime("%H%M%S"), ext)
        log("出力先が使用中のため別名で保存します: " + os.path.basename(alt))
        wb.save(alt)
        out_path = alt
    log("保存しました: " + out_path)
    return out_path


# ---------------------------------------------------------------- 変換本体
def convert(mgt_path, anl_path, out_path=None,
            split_at_supports=True, split_at_releases=True,
            convert_units=True, log=_print):
    if out_path is None:
        out_path = os.path.splitext(mgt_path)[0] + "_断面算定データ.xlsx"
    model = parse_mgt(mgt_path, log)
    results = parse_anl(anl_path, log)

    n_no_force = sum(1 for e in model["elements"].values()
                     if e["id"] not in results["beam_forces"]
                     and e["id"] not in results["truss_forces"])
    if n_no_force:
        log("警告: ANL に断面力が無い要素が {} 本あります(断面力欄は空になります)".format(n_no_force))

    groups = merge_members(model, split_at_supports, split_at_releases, log)
    rows, elem_to_member, units = build_members(model, results, groups,
                                                convert_units, log)
    return write_excel(out_path, model, results, rows, elem_to_member, units, log)


# ---------------------------------------------------------------- GUI
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("MGT/ANL → 断面算定Excel 変換  v" + VERSION)
    root.geometry("640x480")

    frm = ttk.Frame(root, padding=10)
    frm.pack(fill="both", expand=True)

    v_mgt, v_anl, v_out = tk.StringVar(), tk.StringVar(), tk.StringVar()
    v_sup = tk.BooleanVar(value=True)
    v_rls = tk.BooleanVar(value=True)
    v_cnv = tk.BooleanVar(value=True)

    def sync_paths(force=False):
        """MGT に合わせて同名の ANL と出力先をセットする。

        force=True(MGT を選び直したとき)は既存の内容ごと上書きする。
        別モデルの ANL が残ったまま変換される事故を防ぐため、
        同名の ANL が無ければ空欄に戻す。
        """
        p = v_mgt.get()
        if not p:
            return
        stem = os.path.splitext(p)[0]
        if force or not v_anl.get():
            v_anl.set(stem + ".anl" if os.path.exists(stem + ".anl") else "")
        if force or not v_out.get():
            v_out.set(stem + "_断面算定データ.xlsx")

    def picker(var, label, patterns, save=False):
        def _pick():
            if save:
                p = filedialog.asksaveasfilename(
                    title=label, defaultextension=".xlsx", filetypes=patterns)
            else:
                p = filedialog.askopenfilename(title=label, filetypes=patterns)
            if p:
                var.set(p.replace("/", os.sep))
                sync_paths(force=(var is v_mgt))
        return _pick

    rows_def = [
        ("MGT ファイル", v_mgt, [("MGT", "*.mgt"), ("すべて", "*.*")], False),
        ("ANL ファイル", v_anl, [("ANL", "*.anl"), ("すべて", "*.*")], False),
        ("出力先 Excel", v_out, [("Excel", "*.xlsx")], True),
    ]
    for i, (label, var, pats, save) in enumerate(rows_def):
        ttk.Label(frm, text=label).grid(row=i, column=0, sticky="w", pady=3)
        ttk.Entry(frm, textvariable=var).grid(row=i, column=1, sticky="ew",
                                              padx=6, pady=3)
        ttk.Button(frm, text="参照...",
                   command=picker(var, label, pats, save)).grid(row=i, column=2)
    frm.columnconfigure(1, weight=1)

    opt = ttk.LabelFrame(frm, text="部材統合オプション", padding=6)
    opt.grid(row=3, column=0, columnspan=3, sticky="ew", pady=8)
    ttk.Checkbutton(opt, text="支点の節点で部材を切る",
                    variable=v_sup).pack(anchor="w")
    ttk.Checkbutton(opt, text="ピン(梁端解放)の節点で部材を切る",
                    variable=v_rls).pack(anchor="w")
    ttk.Checkbutton(opt, text="断面力・長さを kN・kN·m・mm に換算する",
                    variable=v_cnv).pack(anchor="w")

    btn = ttk.Button(frm, text="変換実行")
    btn.grid(row=4, column=0, columnspan=3, pady=6)

    txt = tk.Text(frm, height=12, state="disabled", font=("MS Gothic", 9))
    txt.grid(row=5, column=0, columnspan=3, sticky="nsew")
    frm.rowconfigure(5, weight=1)

    def log(msg):
        def _append():
            txt.configure(state="normal")
            txt.insert("end", str(msg) + "\n")
            txt.see("end")
            txt.configure(state="disabled")
        root.after(0, _append)

    def run():
        mgt, anl = v_mgt.get().strip(), v_anl.get().strip()
        if not os.path.isfile(mgt):
            messagebox.showerror("エラー", "MGT ファイルを指定してください")
            return
        if not os.path.isfile(anl):
            messagebox.showerror("エラー", "ANL ファイルを指定してください")
            return
        btn.configure(state="disabled")

        def worker():
            try:
                out = convert(mgt, anl, v_out.get().strip() or None,
                              v_sup.get(), v_rls.get(), v_cnv.get(), log)
                log("---- 完了 ----")
                root.after(0, lambda: messagebox.showinfo(
                    "完了", "変換が完了しました。\n" + out))
            except Exception as ex:
                # ex は except ブロックを抜けると消えるため、遅延実行の
                # ラムダに渡す前に文字列へ束縛しておく
                msg = str(ex)
                log("エラー: " + msg)
                log(traceback.format_exc())
                root.after(0, lambda: messagebox.showerror("エラー", msg))
            finally:
                root.after(0, lambda: btn.configure(state="normal"))
        threading.Thread(target=worker, daemon=True).start()

    btn.configure(command=run)
    root.mainloop()


# ---------------------------------------------------------------- main
def main():
    if len(sys.argv) == 1:
        run_gui()
        return
    ap = argparse.ArgumentParser(description="midas iGen MGT/ANL -> 断面算定用 Excel")
    ap.add_argument("mgt", help="MGT ファイル")
    ap.add_argument("anl", nargs="?", help="ANL ファイル(省略時は MGT と同名)")
    ap.add_argument("-o", "--out", help="出力 xlsx パス")
    ap.add_argument("--no-support-split", action="store_true",
                    help="支点で部材を切らない")
    ap.add_argument("--no-release-split", action="store_true",
                    help="ピン(梁端解放)で部材を切らない")
    ap.add_argument("--no-unit-convert", action="store_true",
                    help="単位換算せず元単位のまま出力")
    a = ap.parse_args()
    anl = a.anl or os.path.splitext(a.mgt)[0] + ".anl"
    convert(a.mgt, anl, a.out,
            not a.no_support_split, not a.no_release_split, not a.no_unit_convert)


if __name__ == "__main__":
    main()
